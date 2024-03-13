
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time

class seleniumClass:
    def selens():
        driver_path = "./chromedriver.exe"

        options = webdriver.ChromeOptions()


        driver = webdriver.Chrome(options=options,service=Service(driver_path) )

        driver.get("https://www.facebook.com/login/?privacy_mutation_token=eyJ0eXBlIjowLCJjcmVhdGlvbl90aW1lIjoxNzA4NjUxMDU3LCJjYWxsc2l0ZV9pZCI6MjY5NTQ4NDUzMDcyMDk1MX0%3D")

        driver.find_element(by="id", value="email").send_keys("luquinhas19221@hotmail.com")
        # Encontra o elemento pelo XPath
        driver.find_element(by="id", value="pass").send_keys("ps2pcwii")

        driver.find_element(by="id", value="loginbutton").click()



        workbook = openpyxl.load_workbook('./cadastro2cx.xlsx')
        sheet = workbook['Planilha1']  # Substitua 'Planilha1' pelo nome da sua planilha

        # Ler dados da célula
        nomes = []

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=0, max_col=0):
            for cell in row:
                if cell.value is not None:
                    nomes.append(cell.value)
        print(nomes)
        # while True:
        #     entrada = input("Digite 'q' para fechar o navegador: ")
        #     if entrada.lower() == 'q':
        #         break
        driver.quit()

        return nomes